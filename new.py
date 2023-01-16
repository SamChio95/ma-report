import openpyxl
import glob


def check_model(txt):
    global model
    for line in txt:
        if "ArubaOS (MODEL: " in line and "Aruba7030" not in line:
            model = "other_aruba_model"
        if "Cisco IOS Software, ISR Software" in line:
            model = "cisco_isr"
        if "ArubaOS (MODEL: Aruba7030)" in line:
            model = "aruba7030"
        if "keyword to distinguish cisco switch" in line:
            model = "cisco_switch"

def select_keyword(model):
    global temp_keyword_1
    global temp_keyword_2
    global temp_keyword_3
    global sn_keyword_1
    global total_mem_keyword_1
    global total_mem_keyword_2
    global free_mem_keyword_1
    global free_mem_keyword_2
    if model == "other_aruba_model":
        temp_keyword_1 = "Main Board Temperatures"
        temp_keyword_2 = "die temp"
        temp_keyword_3 = "Temp"
        sn_keyword_1 = "System Serial#            	:"
        total_mem_keyword_1 = "show memory"
        total_mem_keyword_2 = "Memory (Kb): total: "
        free_mem_keyword_1 = "show memory"
        free_mem_keyword_2 = "free: "
    if model == "aruba7030":
        temp_keyword_1 = "Main Board Temperatures"
        temp_keyword_2 = "PHY 0 Temp"
        sn_keyword_1 = "System Serial#            	:" 
        total_mem_keyword_1 = "show memory"
        total_mem_keyword_2 = "Memory (Kb): total: "
        free_mem_keyword_1 = "show memory"
        free_mem_keyword_2 = "free: "        
    if model == "cisco_isr":     
        temp_keyword_1 = "Temp: core-A"
        temp_keyword_2 = "Celsius"
        sn_keyword_1 = "Processor board ID"
    if model == "cisco_switch":
        temp_keyword_1 = "Temp: core-A"
        temp_keyword_2 = "Celsius"
        sn_keyword_1 = "Processor board ID"    

def find_string_after(keyword):
    return line.partition(keyword)[2].strip().partition(" ")[0]

def find_string_before(keyword):
    return line.partition(keyword)[0].strip().rpartition(" ")[-1]

def find_string_after_no_comma(keyword):
    return line.partition(keyword)[2].strip().partition(",")[0]

def search_with_3_keyword_and_output(keyword_1, keyword_2, keyword_3):
    global keyword_1_first_found
    if keyword_1 in line and keyword_1_first_found == "no":
        keyword_1_first_found = "yes"
    if keyword_2 in line and keyword_1_first_found == "yes":
        output = find_string_after(keyword_3)
        active_sheet.cell(excel_start_rows,column_num).value = output
        keyword_1_first_found = "finished"
    
def search_with_2_keyword_and_output(keyword_1, keyword_2):
    global keyword_1_first_found
    if keyword_1 in line and keyword_1_first_found == "no":
        keyword_1_first_found = "yes"
    if keyword_2 in line and keyword_1_first_found == "yes":
        output = find_string_after(keyword_2)
        active_sheet.cell(excel_start_rows,column_num).value = output
        keyword_1_first_found = "finished"

def search_with_2_keyword_reverse_and_output(keyword_1, keyword_2):
    global keyword_1_first_found
    if keyword_1 in line and keyword_1_first_found == "no":
        keyword_1_first_found = "yes"
    if keyword_2 in line and keyword_1_first_found == "yes":
        output = find_string_before(keyword_2)
        active_sheet.cell(excel_start_rows,column_num).value = output
        keyword_1_first_found = "finished"

def search_with_1_keyword_and_output(keyword_1):
    global keyword_1_first_found
    if keyword_1 in line:
        output = find_string_after(keyword_1)
        active_sheet.cell(excel_start_rows,column_num).value = output

def search_aruba_total_mem(keyword_1, keyword_2):
    global keyword_1_first_found
    global total_mem
    if keyword_1 in line and keyword_1_first_found == "no":
        keyword_1_first_found = "yes"
    if keyword_2 in line and keyword_1_first_found == "yes":
        total_mem = float(find_string_after_no_comma(keyword_2))
        keyword_1_first_found = "finished"

def search_aruba_free_mem_and_output(keyword_1, keyword_2):
    global keyword_1_first_found
    free_mem = 1
    if keyword_1 in line and keyword_1_first_found == "no":
        keyword_1_first_found = "yes"
    if keyword_2 in line and keyword_1_first_found == "yes":
        free_mem = float(find_string_after_no_comma(keyword_2))
        output = free_mem/total_mem * 100
        output = "{:.2f}".format(output) + "%"
        active_sheet.cell(excel_start_rows,column_num).value = output
        keyword_1_first_found = "finished"




wb = openpyxl.load_workbook('template.xlsx')
active_sheet = wb.active

excel_start_rows = 2
output = ""
keyword_1_first_found = "no"
free_mem = 1
total_mem = 1

dir_list = glob.glob("*.txt")

for txtfile in dir_list:
    with open(txtfile,'r') as logfile:
        check_model(logfile)
        
    with open(txtfile,'r') as logfile:
        for line in logfile:
            column_num = 2      #serial_number
            select_keyword(model)
            search_with_1_keyword_and_output(sn_keyword_1)
        keyword_1_first_found = "no"

    with open(txtfile,'r') as logfile:
        for line in logfile:            
            column_num = 3      #temperature
            if model == "aruba7030":
                search_with_2_keyword_and_output(temp_keyword_1, temp_keyword_2)
            if model == "cisco_switch":
                search_with_2_keyword_reverse_and_output(temp_keyword_1, temp_keyword_2)
            if model == "cisco_isr":
                search_with_2_keyword_reverse_and_output(temp_keyword_1, temp_keyword_2)
            if model == "other_aruba_model":
                search_with_3_keyword_and_output(temp_keyword_1, temp_keyword_2, temp_keyword_3)
        keyword_1_first_found = "no"

    with open(txtfile,'r') as logfile:
        for line in logfile:            
            if model == "other_aruba_model":
                search_aruba_total_mem(total_mem_keyword_1, total_mem_keyword_2)
        keyword_1_first_found = "no"

    with open(txtfile,'r') as logfile:
        for line in logfile:            
            column_num = 4      #memory
            if model == "other_aruba_model" or "aruba7030":
                search_aruba_free_mem_and_output(free_mem_keyword_1, free_mem_keyword_2)
        keyword_1_first_found = "no"

    excel_start_rows += 1


wb.save('report.xlsx')